"""Build a Sales Order Pick List (xlsx) from extracted LPO JSON.
Reusable by the dynamic tool. Matches the IMG_7355 picklist template layout."""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

ARIAL="Arial"
thin=Side(style="thin",color="BBBBBB")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
HDR=PatternFill("solid",fgColor="1B3A6B")
BOX=PatternFill("solid",fgColor="F0F5FF")

def _set(c,v,*,bold=False,size=10,color="000000",fill=None,align="left",bdr=False):
    c.value=v; c.font=Font(name=ARIAL,bold=bold,size=size,color=color)
    c.alignment=Alignment(horizontal=align,vertical="center")
    if fill:c.fill=fill
    if bdr:c.border=BORDER

def build(data, out_xlsx):
    items=data["items"]
    wb=Workbook(); ws=wb.active; ws.title="Pick List"; ws.sheet_view.showGridLines=False
    for col,w in {"A":5,"B":16,"C":42,"D":8,"E":9,"F":10,"G":9}.items():
        ws.column_dimensions[col].width=w
    ws.merge_cells("A1:G1")
    _set(ws["A1"],"SALES ORDER - PICK LIST",bold=True,size=14,align="center")
    ws.row_dimensions[1].height=24
    # left info box
    left=[(data.get("customer","") or "Customer", True),
          ("Order Type: "+data.get("order_type",""),False),
          ("Vendor: "+data.get("vendor",""),False),
          ("Vendor TRN: "+data.get("vendor_trn",""),False),
          ("Currency: "+data.get("currency","")+"        Ex Rate: 1.0000",False)]
    for r in range(3,8):
        for col in "ABC": ws[f"{col}{r}"].fill=BOX
    for i,(txt,b) in enumerate(left):
        ws.merge_cells(f"A{3+i}:C{3+i}")
        _set(ws[f"A{3+i}"],txt,bold=b,size=(10 if b else 9),color=("1B3A6B" if b else "333333"))
    right=[("Order Date",data.get("order_date","")),
           ("Source File",data.get("source_file","")),
           ("Total Lines",str(len(items))),
           ("Page","Page 1 of 1")]
    for i,(lbl,val) in enumerate(right):
        r=3+i; ws.merge_cells(f"E{r}:G{r}")
        cell=ws[f"E{r}"]
        cell.value=CellRichText(
            TextBlock(InlineFont(rFont=ARIAL,b=True,sz=9,color="1B3A6B"),f"{lbl}:  "),
            TextBlock(InlineFont(rFont=ARIAL,sz=9,color="333333"),str(val)))
        cell.alignment=Alignment(horizontal="left",vertical="center")
    # table header
    heads=["SI","Code (GTIN)","Description","UDM","Checked","Quantity","Picked"]
    for j,h in enumerate(heads):
        _set(ws.cell(9,j+1),h,bold=True,size=10,color="FFFFFF",fill=HDR,align="center",bdr=True)
    ws.row_dimensions[9].height=20
    start=10
    for k,it in enumerate(items):
        r=start+k
        _set(ws.cell(r,1),k+1,align="center",bdr=True,size=9)
        _set(ws.cell(r,2),it.get("gtin",""),align="left",bdr=True,size=9)
        _set(ws.cell(r,3),it.get("desc",""),align="left",bdr=True,size=9)
        _set(ws.cell(r,4),it.get("uom","NOS"),align="center",bdr=True,size=9)
        _set(ws.cell(r,5),"",align="center",bdr=True)
        _set(ws.cell(r,6),it.get("qty",""),align="center",bdr=True,size=9)
        _set(ws.cell(r,7),"",align="center",bdr=True)
    last=start+len(items)-1
    tr=last+1
    _set(ws.cell(tr,3),"Total Quantity:",bold=True,align="right",size=9)
    ws.cell(tr,6).value=f"=SUM(F{start}:F{last})"
    ws.cell(tr,6).font=Font(name=ARIAL,bold=True,size=9)
    ws.cell(tr,6).alignment=Alignment(horizontal="center"); ws.cell(tr,6).border=BORDER
    fr=tr+2
    _set(ws.cell(fr,1),"Prepared By: ______________",size=9); ws.merge_cells(f"A{fr}:C{fr}")
    _set(ws.cell(fr,4),"Checked By: ______________",size=9); ws.merge_cells(f"D{fr}:E{fr}")
    _set(ws.cell(fr,6),"Picked By: ______________",size=9); ws.merge_cells(f"F{fr}:G{fr}")
    ws.print_area=f"A1:G{fr}"
    ws.page_setup.orientation="portrait"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    wb.save(out_xlsx)
    return out_xlsx

if __name__=="__main__":
    data=json.load(open(sys.argv[1],encoding="utf-8"))
    print(build(data, sys.argv[2]))
